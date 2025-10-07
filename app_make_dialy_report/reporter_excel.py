"""
Excel Generator para APP MAKE DAILY REPORT.

Genera archivos Excel formateados con datos de alertas.

Responsabilidades:
- Crear archivos Excel desde diccionarios
- Aplicar formato profesional
- Autoajustar columnas

Autor: Sistema de Tracking Dropi-Inter
Fecha: Octubre 2025
"""

from __future__ import annotations
import os
import logging
from typing import List, Dict, Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment


class ExcelGenerator:
    """
    Generador de archivos Excel formateados.
    """
    
    def generate(
        self,
        data: List[Dict[str, Any]],
        output_dir: str,
        filename: str
    ) -> str:
        """
        Genera archivo Excel con formato.
        
        Args:
            data: Lista de diccionarios con datos
            output_dir: Directorio de salida
            filename: Nombre del archivo
            
        Returns:
            str: Ruta completa del archivo generado
        """
        if not data:
            logging.warning("No hay datos para generar Excel")
            return ""
        
        # Crear DataFrame
        df = pd.DataFrame(data)
        
        # Ruta de salida
        output_path = os.path.join(output_dir, filename)
        
        # Guardar Excel
        logging.info(f"Generando Excel: {output_path}")
        df.to_excel(output_path, index=False, engine='openpyxl')
        
        # Aplicar formato
        self._apply_formatting(output_path)
        
        logging.info(f"Excel generado exitosamente: {len(data)} registros")
        return output_path
    
    @staticmethod
    def _apply_formatting(file_path: str) -> None:
        """
        Aplica formato profesional al archivo Excel.
        
        Args:
            file_path: Ruta del archivo Excel
        """
        wb = load_workbook(file_path)
        ws = wb.active
        
        # Formato de headers
        header_fill = PatternFill(
            start_color="4472C4",
            end_color="4472C4",
            fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Autoajustar columnas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except (TypeError, AttributeError):
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Guardar cambios
        wb.save(file_path)
        logging.info("Formato aplicado al Excel")
